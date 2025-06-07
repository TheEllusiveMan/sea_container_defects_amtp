import os
import io
import time
import shutil
import secrets
from PIL import Image as PILImage
from openpyxl import load_workbook

from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, redirect, url_for, session

from models import *

secret_key = secrets.token_hex(16)

app = Flask(__name__)
app.secret_key = secret_key

model_container = ModelContainer()
model_container_number = ModelContainerNumber()
model_container_damage = ModelContainerDamage()

dmg_translate = {'Deframe': 'деформация',
                 'Hole': 'пробоина',
                 'Rusty': 'ржавчина',
                 'Dent': 'вмятина',
                 'Scratch': 'царапины'}

wall_dict = {'front': 'на передней стенке ',
             'back': 'на задней стенке ',
             'left': 'на левой стенке ',
             'right': 'на правой стенке '}

UPLOAD_FOLDER = 'static/uploads'
TEMP_IMAGES_DIR = 'temp_images'
REPORTS_DIR = 'reports'
MODELS_DIR = 'models'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMP_IMAGES_DIR'] = TEMP_IMAGES_DIR
app.config['REPORTS_DIR'] = REPORTS_DIR
app.config['SECRET_KEY'] = secret_key

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['TEMP_IMAGES_DIR'], exist_ok=True)
os.makedirs(app.config['REPORTS_DIR'], exist_ok=True)
os.makedirs(MODELS_DIR, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_wall_str(list_result_list_number, side_str):
    wall_types = session.get('wall_types', [])
    if list_result_list_number < len(wall_types):
        side_str = wall_dict.get(wall_types[list_result_list_number], '')
    return side_str


def paint_results_dmg_number(cropped_img, results):
    # Отобразите результаты
    for result_list in results:
        for result in result_list:
            boxes = result.boxes  # Получите ограничивающие рамки
            names = result.names
            for box in boxes:
                x1, y1, x2, y2 = box.xyxy[0]  # Координаты bbox
                conf = box.conf[0]  # Уверенность
                cls = box.cls[0]  # Класс
                class_name = names[int(cls)]

                # Нарисуйте bbox на изображении
                cv2.rectangle(cropped_img, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 5)  # Красный цвет
                cv2.putText(cropped_img, f'Class: {class_name}, Conf: {conf:.2f}', (int(x1), int(y1) - 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

    # st.write("Найденные повреждения для:", im_name)
    return cropped_img


def get_cropped_img_by_xyxys(image_rgb, xyxys):
    xyxy_1 = xyxys[0]
    xyxy = xyxy_1[0]
    x_min, y_min, x_max, y_max = xyxy[0], xyxy[1], xyxy[2], xyxy[3]
    x_min, y_min, x_max, y_max = int(x_min), int(y_min), int(x_max), int(y_max)
    cropped_img = image_rgb[y_min:y_max, x_min:x_max]

    return cropped_img


def container_predict(image_rgb, idx):
    # предсказываем контейнеры
    results_cont = model_container.predict(image_rgb)
    # получаем координаты рамок с контейнерами
    frame, xyxys, confidences, class_ids = model_container.plot_bbboxes(results_cont, image_rgb)

    # Отобразите результаты
    for result in results_cont:
        boxes = result.boxes  # Получите ограничивающие рамки
        for box in boxes:
            x1, y1, x2, y2 = box.xyxy[0]  # Координаты bbox
            conf = box.conf[0]  # Уверенность
            cls = box.cls[0]  # Класс

            # Нарисуйте bbox на изображении
            cv2.rectangle(image_rgb, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 3)  # Красный цвет
            cv2.putText(image_rgb, f'Class: {int(cls)}, Conf: {conf:.2f}', (int(x1), int(y1) + 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

    processed_img_path = os.path.join(app.config['UPLOAD_FOLDER'], f'processed_container_{idx}.jpg')
    cv2.imwrite(processed_img_path, cv2.cvtColor(image_rgb, cv2.COLOR_RGB2BGR))

    return xyxys, processed_img_path


def damage_predict(image, xyxys, idx):
    results_dmg = model_container_damage.predict(frame=image, box=xyxys)

    image_rgb = np.array(image)

    cropped_img = get_cropped_img_by_xyxys(image_rgb, xyxys)
    processed_img = paint_results_dmg_number(cropped_img, results_dmg)

    processed_img_path = os.path.join(app.config['UPLOAD_FOLDER'], f'processed_damage_{idx}.jpg')
    cv2.imwrite(processed_img_path, cv2.cvtColor(processed_img, cv2.COLOR_RGB2BGR))

    return results_dmg, processed_img_path


def find_cont_number(image, xyxys, idx):
    results_number = model_container_number.predict(frame=image, box=xyxys)

    image_rgb = np.array(image)

    cropped_img = get_cropped_img_by_xyxys(image_rgb, xyxys)
    processed_img = paint_results_dmg_number(cropped_img, results_number)

    processed_img_path = os.path.join(app.config['UPLOAD_FOLDER'], f'processed_number_{idx}.jpg')
    cv2.imwrite(processed_img_path, cv2.cvtColor(processed_img, cv2.COLOR_RGB2BGR))

    return results_number, processed_img_path


def find_cont_number_with_manual_wall(idx, image, container_xyxys):
    wall_types = session.get('wall_types', [])
    if idx <= len(wall_types):
        wall_type = wall_types[idx-1]
        if wall_type in ["front", "back"]:
            return find_cont_number(image, container_xyxys, idx)
    return None, None


def recogn_number(cont_number_detections):
    result_number = 'TCLO 531461 4'  # Пока заглушка

    return result_number


def results_to_dict(results):
    """Конвертирует объекты Results в словарь для сериализации"""
    if results is None:
        return None

    result_list = []
    for result in results:
        boxes_data = []
        # print('result', type(result), result)
        for res in result:
            for box in res.boxes:
                boxes_data.append({
                    'xyxy': box.xyxy[0].tolist(),
                    'conf': box.conf[0].item(),
                    'cls': box.cls[0].item()
                })
            result_list.append({
                'names': res.names,
                'boxes': boxes_data
            })
    return result_list


def clean_temp_folder():
    """Удаляет все файлы в указанной папке, очищает поле ввода номера контейнера."""

    for filename in os.listdir(app.config['TEMP_IMAGES_DIR']):
        file_path = os.path.join(app.config['TEMP_IMAGES_DIR'], filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Ошибка при удалении {file_path}: {e}")

    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Ошибка при удалении {file_path}: {e}")


def make_report(results_dmg, result_number):
    cont_result_all = ''

    side_str = ''
    for result_num, result in enumerate(results_dmg):
        side_str = get_wall_str(result_num, side_str)
        names = result[0]['names']
        boxes = result[0]['boxes']

        for box in boxes:
            class_name = names[str(int(box['cls']))]
            side_str += dmg_translate[class_name] + ', '

        cont_result_all += side_str

    # Загружаем шаблон
    AOF_template_path = "docs_templates/AOF_template.xlsx"
    wb = load_workbook(AOF_template_path)
    sheet = wb.active   # Получаем активный лист (предполагаем, что это TDSheet)

    # Записываем данные в указанные ячейки
    # R10C2 строка 10, столбец 2 (B10 в Excel)
    sheet.cell(row=10, column=2, value=result_number)
    # R13C1 строка 13, столбец 1 (A13 в Excel)
    sheet.cell(row=13, column=1, value=cont_result_all)

    filename = os.path.join(app.config['REPORTS_DIR'], f'АОФ {result_number}.xlsx')

    # Сохраняем как новый файл
    wb.save(filename)

    return filename


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Handle file upload
        if 'files' not in request.files:
            return redirect(request.url)

        files = request.files.getlist('files')
        if len(files) == 0 or all(file.filename == '' for file in files):
            return redirect(request.url)

        # Save uploaded files
        uploaded_files = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                uploaded_files.append(filename)

        # Save to temp directory for processing
        temp_files = []
        for filename in uploaded_files:
            src = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            dst = os.path.join(app.config['TEMP_IMAGES_DIR'], filename)
            shutil.copy(src, dst)
            temp_files.append(filename)

        session['uploaded_files'] = uploaded_files
        session['temp_files'] = temp_files

        return redirect(url_for('process'))

    return render_template('index.html')


@app.route('/process', methods=['GET', 'POST'])
def process():
    if request.method == 'POST':
        if 'analyze' in request.form:
            # Get wall types from form
            wall_types = []
            for i in range(1, 4):
                wall_type = request.form.get(f'wall_type_{i}')
                if wall_type:
                    wall_types.append(wall_type)
            session['wall_types'] = wall_types

            # Process images
            uploaded_files = session.get('uploaded_files', [])
            temp_files = session.get('temp_files', [])
            wall_types = session.get('wall_types', [])

            cont_dmg_result = []
            processed_images = []

            for idx, filename in enumerate(uploaded_files, start=1):
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                temp_filepath = os.path.join(app.config['TEMP_IMAGES_DIR'], filename)
                curr_im_filename = temp_files[idx-1]

                image_bbs = cv2.imread(temp_filepath)
                image_rgb = cv2.cvtColor(image_bbs, cv2.COLOR_BGR2RGB)
                image_pil = PILImage.open(filepath)

                # Предсказание контейнера
                container_xyxys, container_img = container_predict(image_rgb, idx)

                # Предсказание повреждений
                result_dmg, damage_img = damage_predict(image_pil, container_xyxys, idx)

                # Предсказание маркировки
                cont_number_detections, number_img = find_cont_number_with_manual_wall(idx, image_pil, container_xyxys)

                # Распознавание номера
                result_number = recogn_number(cont_number_detections)

                cont_dmg_result.append(results_to_dict(result_dmg))
                processed_images.append({
                    'container': 'processed_container_' + str(idx) + '.jpg',
                    'damage': 'processed_damage_' + str(idx) + '.jpg',
                    'number': 'processed_number_' + str(idx) + '.jpg' if number_img else None
                })

            session['result_dmg'] = cont_dmg_result
            session['result_number'] = result_number
            session['processed_images'] = processed_images

            return redirect(url_for('results'))

        elif 'report' in request.form:
            user_number_input = request.form.get('container_number', '').strip()
            if user_number_input:
                res_number = user_number_input.upper()
            else:
                res_number = session.get('result_number', 'UNKNOWN')

            res_dmg = session.get('result_dmg', [])
            report_filename = make_report(res_dmg, res_number)

            clean_temp_folder()

            session.pop('uploaded_files', None)
            session.pop('temp_files', None)
            session.pop('result_dmg', None)
            session.pop('result_number', None)
            session.pop('processed_images', None)

            return redirect(url_for('index'))

    return render_template('process.html')

@app.route('/results')
def results():
    processed_images = session.get('processed_images', [])
    result_number = session.get('result_number', 'UNKNOWN')

    return render_template('results.html',
                           images=processed_images,
                           result_number=result_number)


if __name__ == '__main__':
    app.run(debug=True)
